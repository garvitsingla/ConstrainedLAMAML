import os
import time
import random
import builtins
import io
import warnings
import logging
import itertools

import torch
import numpy as np
from collections import OrderedDict
from contextlib import contextmanager, redirect_stdout, redirect_stderr
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

warnings.filterwarnings("ignore")
logging.getLogger("gymnasium").setLevel(logging.ERROR)

from environment import (
    HAZARD_TYPES,
    ConstrainedGoToLocalEnv, ConstrainedPickupDistEnv,
    ConstrainedGoToObjDoorEnv, ConstrainedOpenDoorEnv,
    ConstrainedOpenDoorLocEnv, ConstrainedOpenDoorsOrderEnv,
    ConstrainedActionObjDoorEnv, ConstrainedGoToOpenEnv,
    ConstrainedFindObjS5Env,
)
from sampler_lang import (
    BabyAIMissionTaskWrapper, SentenceMissionEncoder,
    MissionParamAdapter, ConstraintParamAdapter,
    preprocess_obs,
)
from maml_rl.policies.categorical_mlp import CategoricalMLPPolicy

import argparse

# ─────────────────────────────────────────────────────────────────────────────
# Argparser
# ─────────────────────────────────────────────────────────────────────────────
p = argparse.ArgumentParser()
p.add_argument("--env", dest="env_name",
               choices=["ConstrainedGoToLocal","ConstrainedPickupDist","ConstrainedGoToObjDoor",
                        "ConstrainedOpenDoor","ConstrainedOpenDoorLoc","ConstrainedOpenDoorsOrder",
                        "ConstrainedActionObjDoor","ConstrainedGoToOpen","ConstrainedFindObjS5"],
               default="ConstrainedGoToLocal")
p.add_argument("--room-size", type=int, default=8)
p.add_argument("--num-dists", type=int, default=2)
p.add_argument("--max-steps", type=int, default=300)
p.add_argument("--n-missions", type=int, default=10)
p.add_argument("--n-episodes", type=int, default=10)
p.add_argument("--num-constraints", type=int, default=2)
p.add_argument("--deltas", type=float, nargs="+", default=[0.3, 0.5, 0.7, 0.9])
args = p.parse_args()

# ─────────────────────────────────────────────────────────────────────────────
# Seed
# ─────────────────────────────────────────────────────────────────────────────
seed = 42
random.seed(seed)
np.random.seed(seed)
torch.manual_seed(seed)
if torch.cuda.is_available():
    torch.cuda.manual_seed_all(seed)

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")


# ─────────────────────────────────────────────────────────────────────────────
# Silence helper
# ─────────────────────────────────────────────────────────────────────────────
@contextmanager
def silence():
    real_print = builtins.print
    buf = io.StringIO()
    def fp(*args, **kwargs):
        if args and isinstance(args[0], str) and "Sampling rejected" in args[0]:
            return
        real_print(*args, **kwargs)
    builtins.print = fp
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            yield
    finally:
        builtins.print = real_print


# ─────────────────────────────────────────────────────────────────────────────
# File Lock helper (to match evaluate_summary.py)
# ─────────────────────────────────────────────────────────────────────────────
@contextmanager
def file_lock(lock_path, timeout=10):
    lock_file = lock_path + ".lock"
    start_time = time.time()
    while True:
        try:
            fd = os.open(lock_file, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            break
        except FileExistsError:
            if time.time() - start_time > timeout:
                raise TimeoutError(f"Could not acquire lock on {lock_path} within {timeout} seconds.")
            time.sleep(0.5)
    try:
        yield
    finally:
        try:
            os.remove(lock_file)
        except OSError:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# Mission definitions
# ─────────────────────────────────────────────────────────────────────────────
OBJECTS     = ['box']
COLORS      = ['red','green','blue','purple','yellow','grey']
PREP_LOCS   = ['on','at','to']
LOC_NAMES   = ['right','front']
DOOR_COLORS = ['yellow','grey']

CONSTRAINT_TEXTS = [f"avoid {h}" for h in HAZARD_TYPES]
DOUBLE_CONSTRAINT_TEXTS = [
    f"avoid {h1} and avoid {h2}"
    for h1, h2 in itertools.combinations(HAZARD_TYPES.keys(), 2)
]

LOCAL_MISSIONS    = [f"go to the {c} {o}" for c in COLORS for o in OBJECTS]
PICKUP_MISSIONS   = [f"pick up the {c} {o}" for c in COLORS for o in OBJECTS]
DOOR_MISSIONS     = [f"go to the {c} door" for c in DOOR_COLORS]
OPENDOOR_MISSIONS = [f"open the {c} door" for c in DOOR_COLORS]
OPENDOORLOC_MISSIONS = [f"open the door {p} the {l}" for p in PREP_LOCS for l in LOC_NAMES]
OPENDOORSORDER_MISSIONS = (
    [f"open the {c} door" for c in DOOR_COLORS] +
    [f"open the {c1} door, then open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS] +
    [f"open the {c1} door after you open the {c2} door" for c1 in DOOR_COLORS for c2 in DOOR_COLORS]
)
ACTIONOBJDOOR_MISSIONS = (
    [f"pick up the {c} {t}" for c in COLORS for t in ["box"]] +
    [f"go to the {c} {t}" for c in COLORS for t in ["box"]] +
    [f"go to the {c} door" for c in DOOR_COLORS] +
    [f"open the {c} door" for c in DOOR_COLORS]
)
FINDOBJS5_MISSIONS = [f"pick up the {t}" for t in ["box"]]

GOALS_MAP = {
    "ConstrainedGoToLocal":      LOCAL_MISSIONS,
    "ConstrainedPickupDist":     PICKUP_MISSIONS,
    "ConstrainedGoToObjDoor":    LOCAL_MISSIONS + DOOR_MISSIONS,
    "ConstrainedGoToOpen":       LOCAL_MISSIONS,
    "ConstrainedOpenDoor":       OPENDOOR_MISSIONS,
    "ConstrainedOpenDoorLoc":    OPENDOOR_MISSIONS + OPENDOORLOC_MISSIONS,
    "ConstrainedOpenDoorsOrder": OPENDOORSORDER_MISSIONS,
    "ConstrainedActionObjDoor":  ACTIONOBJDOOR_MISSIONS,
    "ConstrainedFindObjS5":      FINDOBJS5_MISSIONS,
}


# ─────────────────────────────────────────────────────────────────────────────
# Environment builder
# ─────────────────────────────────────────────────────────────────────────────
def build_env(env_name, room_size, num_dists, max_steps, missions, goals, constraints):
    room_size = room_size if room_size != "env" else args.room_size
    num_dists = num_dists if num_dists != "env" else args.num_dists

    dispatch = {
        "ConstrainedGoToLocal":      lambda: ConstrainedGoToLocalEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps),
        "ConstrainedPickupDist":     lambda: ConstrainedPickupDistEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps),
        "ConstrainedGoToObjDoor":    lambda: ConstrainedGoToObjDoorEnv(max_steps=max_steps, num_distractors=num_dists),
        "ConstrainedGoToOpen":       lambda: ConstrainedGoToOpenEnv(room_size=room_size, num_dists=num_dists, max_steps=max_steps),
        "ConstrainedOpenDoor":       lambda: ConstrainedOpenDoorEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedOpenDoorLoc":    lambda: ConstrainedOpenDoorLocEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedOpenDoorsOrder": lambda: ConstrainedOpenDoorsOrderEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedActionObjDoor":  lambda: ConstrainedActionObjDoorEnv(room_size=room_size, max_steps=max_steps),
        "ConstrainedFindObjS5":      lambda: ConstrainedFindObjS5Env(room_size=5, max_steps=max_steps),
    }
    base = dispatch[env_name]()
    return BabyAIMissionTaskWrapper(base, missions=missions, goals=goals, constraints=constraints)


# ─────────────────────────────────────────────────────────────────────────────
# Evaluation configs per environment
# ─────────────────────────────────────────────────────────────────────────────
def get_configs(env_name):
    if env_name in ["ConstrainedGoToLocal", "ConstrainedPickupDist"]:
        return [(7, 3), (7, 5), (8, 2), (8, 4), (9, 3), (9, 5)]
    elif env_name == "ConstrainedGoToObjDoor":
        return [("env", 1), ("env", 2), ("env", 3), ("env", 4), ("env", 5)]
    elif env_name == "ConstrainedActionObjDoor":
        return [("env", "env")]
    elif env_name in ["ConstrainedGoToOpen", "ConstrainedFindObjS5"]:
        return [(5, 2), (5, 3), (6, 2), (6, 4)]
    else:
        return [(6, "env"), (7, "env"), (8, "env"), (9, "env"), (10, "env")]


# ─────────────────────────────────────────────────────────────────────────────
# Single-episode rollout
# ─────────────────────────────────────────────────────────────────────────────
def evaluate_policy(env, policy, params=None, seed=None):
    with silence():
        obs, _ = env.reset(seed=seed)
    done, steps, success, viols = False, 0, False, 0
    env_max = getattr(env.unwrapped, 'max_steps', args.max_steps)

    while not done and steps < env_max:
        obs_t = torch.from_numpy(preprocess_obs(obs)[None]).float().to(device)
        with torch.no_grad():
            if params is not None:
                action = policy(obs_t, params=params).sample().item()
            else:
                action = policy(obs_t).sample().item()
        obs, reward, terminated, truncated, info = env.step(action)
        done = terminated or truncated
        steps += 1
        viols += int(info.get('cost', 0) > 0)
        if terminated:
            success = True

    return steps, success, viols


# ─────────────────────────────────────────────────────────────────────────────
# Adapted params for unified LA-MAML
# ─────────────────────────────────────────────────────────────────────────────
def get_adapted_params(mission, policy, encoder, adapter, delta_theta):
    combined = f"{mission[0]} and {mission[1]}" if isinstance(mission, tuple) else mission
    with torch.no_grad():
        emb = encoder(combined).to(device)
        deltas = adapter(emb)
        names = list(dict(policy.named_parameters()).keys())
        params = list(policy.parameters())
        return OrderedDict(
            (n, p + d.squeeze(0) * delta_theta)
            for n, p, d in zip(names, params, deltas)
        )


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
env_name    = args.env_name
max_steps   = args.max_steps
n_missions  = args.n_missions
n_episodes  = args.n_episodes
nc          = args.num_constraints

# Build mission lists
goals_list = GOALS_MAP[env_name]
constraints_list = CONSTRAINT_TEXTS if nc == 1 else DOUBLE_CONSTRAINT_TEXTS
all_missions = [f"{g} and {c}" for g in goals_list for c in constraints_list]

# Create dummy env for shapes
dummy_env = build_env(env_name, args.room_size, args.num_dists, max_steps,
                      all_missions, goals_list, constraints_list)
dummy_obs, _ = dummy_env.reset()
input_size   = preprocess_obs(dummy_obs).shape[0]
output_size  = dummy_env.action_space.n
hidden_sizes = (64, 64)
nonlinearity = torch.nn.functional.tanh

def make_policy():
    return CategoricalMLPPolicy(
        input_size=input_size, output_size=output_size,
        hidden_sizes=hidden_sizes, nonlinearity=nonlinearity,
    ).to(device)

policy_param_shapes = [p.shape for p in make_policy().parameters()]

encoder = SentenceMissionEncoder(
    model_name="all-MiniLM-L6-v2", frozen=True,
    normalize=True, cache=True, device=device,
)
encoder.eval()
enc_dim = encoder.output_dim

# Load available checkpoints
available_deltas = []
policies = {}
adapters = {}

print(f"\nSearching for Unified LA-MAML checkpoints...")
for dt in args.deltas:
    dt_str = str(int(dt)) if dt == int(dt) else str(dt)
    ckpt_path = f"unified_model/lang_{env_name}_dt{dt_str}_{nc}c.pth"
    if not os.path.exists(ckpt_path):
        # Fallback to standard float format string
        ckpt_path = f"unified_model/lang_{env_name}_dt{dt}_{nc}c.pth"
        
    if os.path.exists(ckpt_path):
        print(f"  [✓] Found checkpoint for delta={dt}: {ckpt_path}")
        ckpt = torch.load(ckpt_path, map_location=device)
        
        # Policy
        pol = make_policy()
        pol.load_state_dict(ckpt["policy"])
        pol.eval()
        policies[dt] = pol
        
        # Adapter
        ad = MissionParamAdapter(enc_dim, policy_param_shapes).to(device)
        ad.load_state_dict(ckpt["mission_adapter"])
        ad.eval()
        adapters[dt] = ad
        
        available_deltas.append(dt)
    else:
        print(f"  [✗] Checkpoint not found for delta={dt}: {ckpt_path}")

if not available_deltas:
    raise RuntimeError(f"No Unified LA-MAML checkpoints found for {env_name} ({nc}c)!")

# ─────────────────────────────────────────────────────────────────────────────
# Evaluation loop (Averaged over all configurations)
# ─────────────────────────────────────────────────────────────────────────────
configs    = get_configs(env_name)
test_tasks = random.sample(all_missions, min(n_missions, len(all_missions)))

print(f"\n{'='*70}")
print(f"Evaluating Unified LA-MAML: {env_name}")
print(f"Tasks: {n_missions} | Episodes/task: {n_episodes} | Deltas: {available_deltas}")
print(f"{'='*70}\n")

# Process one delta at a time
for dt in available_deltas:
    print(f"Evaluating Delta Theta = {dt}...")
    results = {'steps': [], 'successes': [], 'viols': []}

    for c_room, c_dists in configs:
        env = build_env(env_name, c_room, c_dists, max_steps,
                        all_missions, goals_list, constraints_list)

        for mission in test_tasks:
            ep_seeds = [random.randint(0, 1_000_000) for _ in range(n_episodes)]
            theta = get_adapted_params(mission, policies[dt], encoder, adapters[dt], dt)
            
            for ep in range(n_episodes):
                env.reset_task(mission)
                s, ok, v = evaluate_policy(env, policies[dt], params=theta, seed=ep_seeds[ep])
                results['steps'].append(s)
                results['successes'].append(ok)
                results['viols'].append(v)

    # Compute overall statistics for this delta
    mean_steps = np.mean(results['steps']) if results['steps'] else 0.0
    std_steps  = np.std(results['steps'])  if results['steps'] else 0.0
    mean_sr    = round(np.mean(results['successes']), 2) if results['successes'] else 0.0
    mean_viols = np.mean(results['viols']) if results['viols'] else 0.0
    std_viols  = np.std(results['viols'])  if results['viols'] else 0.0

    print(f"  Delta={dt:3.1f}:  SR={mean_sr*100:.2f}%  "
          f"Steps={mean_steps:.2f} ± {std_steps:.2f}  "
          f"Viols={mean_viols:.2f} ± {std_viols:.2f}")

    # ─────────────────────────────────────────────────────────────────────────
    # Save/Update Excel (with file lock)
    # ─────────────────────────────────────────────────────────────────────────
    xlsx_path = "unified_results.xlsx"
    avg_row = [dt, f"{mean_steps:.2f} ± {std_steps:.2f}", mean_sr, f"{mean_viols:.2f} ± {std_viols:.2f}"]

    with file_lock(xlsx_path):
        if os.path.exists(xlsx_path):
            wb = load_workbook(xlsx_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        sheet_name = f"{env_name}_{nc}c"[:31]
        
        # If it's a new sheet, create it with header. Otherwise, append/overwrite
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            header = ["Delta Theta", "Avg Steps", "Success Prob", "Avg Viols"]
            ws.append(header)
            for cell in ws[1]:
                cell.font = Font(bold=True)
        else:
            ws = wb[sheet_name]

        # Check if row with this delta already exists in column 1
        row_idx = None
        for r in range(2, ws.max_row + 1):
            val_dt = ws.cell(row=r, column=1).value
            try:
                if val_dt is not None:
                    if abs(float(val_dt) - dt) < 1e-5:
                        row_idx = r
                        break
            except (ValueError, TypeError):
                continue

        if row_idx is not None:
            # Overwrite existing row
            for col_idx, val in enumerate(avg_row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            print(f"  Updated entry for delta={dt} in sheet '{sheet_name}'")
        else:
            # Append new row
            ws.append(avg_row)
            print(f"  Appended new entry for delta={dt} in sheet '{sheet_name}'")

        wb.save(xlsx_path)

print(f"\nAll summary results successfully saved to {xlsx_path}")
